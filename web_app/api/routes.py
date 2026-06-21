"""API 路由 - 查询、历史、导出等"""
from flask import Blueprint, jsonify, request, send_file, current_app
from flask_login import login_required, current_user
from models import db, QueryHistory
from datetime import datetime
import json
import sys
from pathlib import Path
import threading
import requests

api_bp = Blueprint('api', __name__)

# coordinator 将在第一次使用时延迟导入
_coordinator = None

def get_coordinator():
    """延迟加载 coordinator"""
    global _coordinator
    if _coordinator is None:
        # 添加项目根目录到路径
        project_root = Path(current_app.config['PROJECT_ROOT'])
        sys.path.insert(0, str(project_root / "agents" / "multi-agent"))
        from coordinator import MultiAgentCoordinator
        _coordinator = MultiAgentCoordinator()
    return _coordinator

# 用于控制并发的信号量
from threading import Semaphore
task_semaphore = Semaphore(3)  # 最多3个并发任务


@api_bp.route('/query', methods=['POST'])
@login_required
def create_query():
    """创建新查询"""
    data = request.get_json()
    query = data.get('query', '').strip()
    platforms = data.get('platforms', [])
    location = (data.get('location') or '').strip()
    sort_order = data.get('sort_order', 'none')
    search_keywords = data.get('search_keywords', [])
    category = data.get('category', '')
    sub_category = data.get('sub_category', '')
    budget_min = data.get('budget_min')
    budget_max = data.get('budget_max')

    try:
        sample_count = int(data.get('sample_count', 50))
    except (TypeError, ValueError):
        sample_count = 50

    try:
        budget_min = float(budget_min) if budget_min not in (None, '', False) else None
    except (TypeError, ValueError):
        budget_min = None

    try:
        budget_max = float(budget_max) if budget_max not in (None, '', False) else None
    except (TypeError, ValueError):
        budget_max = None

    sample_count = max(1, min(sample_count, 500))
    if sort_order not in {'none', 'price_asc', 'price_desc'}:
        sort_order = 'none'

    if not query:
        return jsonify({'error': '请输入搜索关键词'}), 400

    if not platforms:
        return jsonify({'error': '请至少选择一个平台'}), 400

    # 创建查询记录
    query_history = QueryHistory(
        user_id=current_user.id,
        query=query,
        platforms=','.join(platforms),
        location=location,
        sample_count=sample_count,
        sort_order=sort_order,
        crawl_progress=f"0/{sample_count}",
        status='pending'
    )
    db.session.add(query_history)
    db.session.commit()

    # 异步执行查询
    thread = threading.Thread(
        target=execute_query_task,
        args=(
            query_history.id,
            query,
            platforms,
            sample_count,
            location,
            sort_order,
            search_keywords,
            category,
            sub_category,
            budget_min,
            budget_max,
        )
    )
    thread.daemon = True
    thread.start()

    return jsonify({
        'success': True,
        'query_id': query_history.id,
        'message': '查询已提交，正在处理中...'
    })


def execute_query_task(query_id, query, platforms, sample_count=50, location='福建', sort_order='none',
                       search_keywords=None, category='', sub_category='', budget_min=None, budget_max=None):
    """执行查询任务（后台线程）"""
    with task_semaphore:
        from app import app
        with app.app_context():
            query_record = db.session.get(QueryHistory, query_id)
            if not query_record:
                return

            try:
                # 更新状态
                query_record.status = 'processing'
                db.session.commit()

                def update_crawl_progress(progress):
                    record = db.session.get(QueryHistory, query_id)
                    if not record:
                        return
                    estimated_total = progress.get('estimated_total', 0)
                    crawled_count = progress.get('crawled_count', 0)
                    target_count = progress.get('target_count', sample_count)
                    record.estimated_total = estimated_total
                    record.crawl_progress = progress.get(
                        'progress_text',
                        f"{crawled_count}/{estimated_total or target_count}"
                    )
                    db.session.commit()

                # 执行查询
                coordinator = get_coordinator()
                result = coordinator.execute_research(
                    query=query,
                    platforms=platforms,
                    sample_count=sample_count,
                    location=location,
                    sort_order=sort_order,
                    search_keywords=search_keywords,
                    category=category,
                    sub_category=sub_category,
                    budget_min=budget_min,
                    budget_max=budget_max,
                    progress_callback=update_crawl_progress
                )

                # 更新记录
                if result.get('success'):
                    crawl_meta = result.get('crawl_meta') or result.get('results', {}).get('crawl_meta') or {}
                    if result.get('products_count', 0) == 0 and crawl_meta.get('failure_reason'):
                        query_record.status = 'failed'
                        query_record.error_message = crawl_meta.get('failure_reason')
                        query_record.completed_at = datetime.utcnow()
                        query_record.report_path = None
                        query_record.products_count = 0
                        query_record.estimated_total = crawl_meta.get('estimated_total', 0)
                        query_record.crawl_progress = crawl_meta.get('progress_text', f"0/{sample_count}")
                        query_record.result_data = json.dumps({
                            'crawl_meta': crawl_meta
                        }, ensure_ascii=False)
                        return

                    query_record.status = 'completed'
                    query_record.products_count = result.get('products_count', 0)
                    query_record.sample_count = sample_count
                    query_record.location = location
                    query_record.sort_order = sort_order
                    query_record.report_path = result.get('report_path')
                    query_record.elapsed_time = result.get('elapsed_time')

                    # 序列化结果数据，处理 Product 对象
                    results_data = result.get('results', {})
                    serializable_results = {}

                    # 转换 Product 对象列表为字典
                    if 'products' in results_data:
                        serializable_results['products'] = [
                            p.to_dict() for p in results_data['products']
                        ]

                    # 复制其他数据
                    for key in ['review_analysis', 'price_monitoring', 'report_path', 'charts']:
                        if key in results_data:
                            serializable_results[key] = results_data[key]

                    crawl_meta = results_data.get('crawl_meta') or result.get('crawl_meta') or {}
                    if crawl_meta:
                        serializable_results['crawl_meta'] = crawl_meta
                        query_record.estimated_total = crawl_meta.get('estimated_total', 0)
                        query_record.crawl_progress = crawl_meta.get('progress_text', f"{query_record.products_count}/{sample_count}")
                    else:
                        query_record.estimated_total = max(query_record.products_count, sample_count)
                        query_record.crawl_progress = f"{query_record.products_count}/{query_record.estimated_total}"

                    query_record.result_data = json.dumps(serializable_results, ensure_ascii=False)
                    query_record.completed_at = datetime.utcnow()
                else:
                    query_record.status = 'failed'
                    query_record.error_message = result.get('error', '未知错误')
                    query_record.completed_at = datetime.utcnow()
                    query_record.report_path = None

                    crawl_meta = result.get('crawl_meta') or result.get('results', {}).get('crawl_meta') or {}
                    if crawl_meta:
                        query_record.estimated_total = crawl_meta.get('estimated_total', query_record.estimated_total or 0)
                        query_record.crawl_progress = crawl_meta.get(
                            'progress_text',
                            query_record.crawl_progress or f"0/{sample_count}"
                        )
                        query_record.result_data = json.dumps({
                            'crawl_meta': crawl_meta
                        }, ensure_ascii=False)

            except Exception as e:
                query_record.status = 'failed'
                query_record.error_message = str(e)
                query_record.completed_at = datetime.utcnow()
                # 记录详细错误信息到日志
                import traceback
                print(f"[ERROR] Query {query_id} failed: {str(e)}")
                print(traceback.format_exc())

            finally:
                db.session.commit()


@api_bp.route('/query/<int:query_id>', methods=['GET'])
@login_required
def get_query(query_id):
    """获取查询详情"""
    query = db.session.query(QueryHistory).filter_by(
        id=query_id,
        user_id=current_user.id
    ).first()

    if not query:
        return jsonify({'error': '查询不存在'}), 404

    result = query.to_dict()

    # 如果已完成或失败且保存了诊断数据，附加结果数据
    if query.status in {'completed', 'failed'} and query.result_data:
        try:
            result['data'] = json.loads(query.result_data)
        except:
            pass

    return jsonify(result)


@api_bp.route('/history', methods=['GET'])
@login_required
def get_history():
    """获取用户查询历史"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)

    pagination = db.session.query(QueryHistory).filter_by(
        user_id=current_user.id
    ).order_by(
        QueryHistory.created_at.desc()
    ).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    return jsonify({
        'items': [item.to_dict() for item in pagination.items],
        'total': pagination.total,
        'page': page,
        'per_page': per_page,
        'pages': pagination.pages
    })


@api_bp.route('/query/<int:query_id>/report', methods=['GET'])
@login_required
def download_report(query_id):
    """下载报告"""
    query = db.session.query(QueryHistory).filter_by(
        id=query_id,
        user_id=current_user.id
    ).first()

    if not query or not query.report_path:
        return jsonify({'error': '报告不存在'}), 404

    report_path = Path(query.report_path)
    if not report_path.exists():
        return jsonify({'error': '报告文件不存在'}), 404

    return send_file(
        report_path,
        as_attachment=True,
        download_name=f"report_{query.query}_{query.id}.md"
    )


@api_bp.route('/analyze-intent', methods=['POST'])
@login_required
def analyze_intent():
    """AI 解析用户搜索意图"""
    import re
    from shared.utils.api_client import ClaudeAPIClient

    data = request.get_json()
    user_input = data.get('query', '').strip()

    if not user_input or len(user_input) < 2:
        return jsonify({'error': '请输入更详细的描述'}), 400

    try:
        project_root = Path(current_app.config['PROJECT_ROOT'])
        # 直接读取配置文件，不经过 env 覆盖（避免被本地代理劫持）
        import json as _json
        config_path = project_root / "config" / "agent_config.json"
        with open(config_path, "r", encoding="utf-8-sig") as _f:
            raw_config = _json.load(_f)
        api_config = raw_config.get("api", {})

        api_client = ClaudeAPIClient(
            api_key=api_config.get("api_key", ""),
            base_url=api_config.get("base_url", "https://api.deepseek.com/anthropic"),
            model=api_config.get("model", "claude-opus-4-7"),
        )

        system_prompt = """你是一个专业的电商搜索意图解析器。用户用中文自然语言描述想买什么，你要提取高精度的结构化参数，用于驱动多平台商品搜索。

## 输出格式

输出严格的 JSON 对象（不要额外文字），包含：

{
  "category": "商品大类",
  "sub_category": "商品细分类目或 null",
  "brand": "品牌或 null",
  "product_line": "产品系列/型号或 null",
  "condition": "全新/二手/不限",
  "budget_min": 数字或 null,
  "budget_max": 数字或 null,
  "key_specs": {"规格名": "规格值"},
  "preferences": ["偏好1", "偏好2"],
  "exclude": ["排除项1"],
  "use_case": "使用场景或 null",
  "search_keywords": ["关键词变体1", "关键词变体2"],
  "confidence": 0.0-1.0,
  "ambiguity_note": "歧义说明或 null",
  "user_profile_hint": "用户画像推断或 null"
}

## 品类识别（重要：必须细分）

大类 → 细分类目映射：
- 手机 → iPhone/安卓手机/老人机/游戏手机
- 电脑 → 笔记本/台式机/一体机/平板
- 耳机 → 真无线/头戴式/骨传导/游戏耳机
- 相机 → 微单/单反/卡片机/运动相机
- 鞋 → 运动鞋/跑鞋/篮球鞋/板鞋/皮鞋/靴子/拖鞋
- 服装 → T恤/衬衫/卫衣/外套/裤子/裙子
- 自行车 → 山地车/公路车/折叠车/电助力/城市通勤车
- 家具 → 沙发/床/桌/椅/柜/书架
- 家电 → 空调/冰箱/洗衣机/电视/吸尘器
- 美妆 → 护肤品/彩妆/香水/面膜
- 食品 → 水果/零食/饮料/生鲜
- 图书 → 小说/教材/漫画/工具书
- 母婴 → 奶粉/纸尿裤/玩具/童装
- 其他

## 价格识别

支持多种中文表达：
- "100以内" "不超过200" "200以下" → budget_max
- "100以上" "至少150" "大于200" → budget_min
- "100到200" "100-200" "一百到两百" "100左右" → budget_min=100, budget_max=200
- "两三百" → budget_min=200, budget_max=300
- "几百块" → budget_min=100, budget_max=999
- "千元机" "千元左右" → budget_min=800, budget_max=1500
- "一两千" → budget_min=1000, budget_max=2000
- "小几百" → budget_min=100, budget_max=400
- "大几百" → budget_min=600, budget_max=999
- 中文数字转阿拉伯数字：一百=100，两百=200，一千=1000，一万=10000

## 隐性需求推断

根据用户的描述推断隐含需求：
- "冬天穿的" → 保暖、加厚、羽绒
- "送女朋友" → 颜值高、包装好、品牌货
- "学生用" → 性价比、预算敏感
- "办公用" → 稳定、续航好
- "打游戏" → 高性能、高刷新率
- "运动穿" → 透气、轻便
- "自用" + 低预算 → 二手可接受
- "拍照" → 相机像素高、存储大

## 关键词生成

search_keywords 要：
1. 覆盖不同表述方式（全称/简称/昵称/英文）
2. 包含规格组合（"iPhone 256G"、"跑鞋 42码"）
3. 从精准到宽泛排序（前面是精准匹配，后面是泛搜）
4. 至少生成 5 个，最多 10 个

## 示例

输入: "100到200元的篮球鞋，42码"
输出:
{
  "category": "鞋",
  "sub_category": "篮球鞋",
  "brand": null,
  "product_line": null,
  "condition": "不限",
  "budget_min": 100,
  "budget_max": 200,
  "key_specs": {"尺码": "42"},
  "preferences": [],
  "exclude": [],
  "use_case": "运动-篮球",
  "search_keywords": ["篮球鞋 42码", "篮球鞋 42", "男篮球鞋 42码", "实战篮球鞋", "球鞋 42"],
  "confidence": 0.95,
  "ambiguity_note": null,
  "user_profile_hint": "运动爱好者，预算有限"
}

输入: "苹果"
输出:
{
  "category": "手机",
  "sub_category": "iPhone",
  "brand": "Apple",
  "product_line": null,
  "condition": "不限",
  "budget_min": null,
  "budget_max": null,
  "key_specs": {},
  "preferences": [],
  "exclude": ["苹果水果", "苹果食品", "苹果配件非手机"],
  "search_keywords": ["iPhone", "苹果手机", "iPhone 15", "iPhone 14", "苹果 手机"],
  "confidence": 0.6,
  "ambiguity_note": "「苹果」有歧义，可能是水果（红富士苹果）或手机（iPhone）。根据电商购物场景默认理解为苹果手机，如用户想要水果请明确说明。",
  "user_profile_hint": null
}

输入: "想买一辆公路车，入门级，3000以内，最好是捷安特或者美利达"
输出:
{
  "category": "自行车",
  "sub_category": "公路车",
  "brand": null,
  "product_line": "入门级",
  "condition": "不限",
  "budget_min": null,
  "budget_max": 3000,
  "key_specs": {},
  "preferences": ["捷安特", "美利达", "入门级"],
  "exclude": [],
  "search_keywords": ["公路车 捷安特", "公路车 美利达", "入门公路车", "捷安特 公路车 3000以内", "美利达 公路车 入门", "公路自行车 3000", "Giant 公路车", "Merida 公路车"],
  "confidence": 0.92,
  "ambiguity_note": null,
  "user_profile_hint": "骑行入门爱好者，品牌倾向明显"
}"""

        result = api_client.send_message(
            prompt=user_input,
            system_prompt=system_prompt,
            max_tokens=1024,
            temperature=0.3,
        )

        text = result.get('text', '')
        # 尝试提取 JSON
        json_match = re.search(r'\{[\s\S]*\}', text)
        if json_match:
            parsed = json.loads(json_match.group())
        else:
            parsed = json.loads(text)

        return jsonify({'success': True, 'intent': parsed})

    except json.JSONDecodeError:
        return jsonify({'success': False, 'error': 'AI 返回格式解析失败，请重试', 'raw': text[:500]}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': f'意图分析失败：{str(e)}'}), 500


@api_bp.route('/charts/<path:filename>')
@login_required
def serve_chart(filename):
    """提供图表图片"""
    from flask import send_from_directory
    project_root = Path(current_app.config['PROJECT_ROOT'])
    charts_dir = project_root / "reports" / "charts"
    return send_from_directory(charts_dir, filename)


@api_bp.route('/skills', methods=['GET'])
@login_required
def list_skills():
    """列出所有可用的 Agent Skills"""
    from shared.utils.skill_loader import get_skill_loader
    project_root = Path(current_app.config['PROJECT_ROOT'])
    loader = get_skill_loader(project_root / "skills")
    return jsonify({'skills': loader.list_skills()})


@api_bp.route('/skills/<skill_name>', methods=['GET'])
@login_required
def get_skill(skill_name):
    """获取指定 Skill 的完整内容"""
    from shared.utils.skill_loader import get_skill_loader
    project_root = Path(current_app.config['PROJECT_ROOT'])
    loader = get_skill_loader(project_root / "skills")
    skill = loader.get_skill(skill_name)
    if not skill:
        return jsonify({'error': f'Skill {skill_name} 不存在'}), 404
    return jsonify({
        'name': skill.name,
        'description': skill.description,
        'body': skill.body,
        'metadata': skill.metadata,
        'resources': skill.list_resources(),
    })


@api_bp.route('/query/<int:query_id>/excel', methods=['GET'])
@login_required
def download_excel(query_id):
    """导出 Excel 格式的商品数据"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    query = db.session.query(QueryHistory).filter_by(
        id=query_id,
        user_id=current_user.id
    ).first()

    if not query or query.status != 'completed' or not query.result_data:
        return jsonify({'error': '数据不存在'}), 404

    try:
        data = json.loads(query.result_data)
    except Exception:
        return jsonify({'error': '数据解析失败'}), 500

    products = data.get('products', [])

    wb = Workbook()
    ws = wb.active
    ws.title = '商品数据'

    # 表头样式
    header_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D5D8DC'),
        right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'),
        bottom=Side(style='thin', color='D5D8DC'),
    )

    headers = ['序号', '平台', '商品标题', '价格(¥)', '销量', '好评率', '评价数',
               '卖家名称', '卖家信誉', '卖家评分', '商品链接', '抓取时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据行
    data_font = Font(name='Microsoft YaHei', size=10)
    platform_map = {'xianyu': '闲鱼', 'taobao': '淘宝', 'jd': '京东', 'pdd': '拼多多'}

    for i, p in enumerate(products, 1):
        seller = p.get('seller') or {}
        row_data = [
            i,
            platform_map.get(p.get('platform', ''), p.get('platform', '')),
            p.get('title', ''),
            p.get('price', 0),
            p.get('sales', 0),
            f"{(p.get('positive_rate', 0) or 0) * 100:.1f}%",
            p.get('review_count', 0),
            seller.get('name', ''),
            seller.get('reputation_level', ''),
            seller.get('rating', ''),
            p.get('url', ''),
            p.get('crawled_at', ''),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            if col == 4:  # 价格列右对齐
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '¥#,##0.00'
            elif col == 1:
                cell.alignment = Alignment(horizontal='center')

    # 列宽
    col_widths = [6, 8, 45, 12, 10, 10, 10, 20, 12, 10, 40, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 自动筛选
    ws.auto_filter.ref = f'A1:{ws.cell(row=1, column=len(headers)).column_letter}{len(products) + 1}'

    # 写入 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_query = query.query.replace('/', '_').replace('\\', '_')[:30]
    filename = f'{safe_query}_商品数据_{len(products)}条.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@api_bp.route('/stats', methods=['GET'])
@login_required
def get_stats():
    """获取用户统计信息"""
    total_queries = db.session.query(QueryHistory).filter_by(user_id=current_user.id).count()
    completed_queries = db.session.query(QueryHistory).filter_by(
        user_id=current_user.id,
        status='completed'
    ).count()

    return jsonify({
        'total_queries': total_queries,
        'completed_queries': completed_queries,
        'failed_queries': db.session.query(QueryHistory).filter_by(
            user_id=current_user.id,
            status='failed'
        ).count(),
        'processing_queries': db.session.query(QueryHistory).filter_by(
            user_id=current_user.id,
            status='processing'
        ).count()
    })


@api_bp.route('/music/search', methods=['POST'])
@login_required
def search_music():
    """搜索音乐"""
    data = request.get_json()
    query = data.get('query', '').strip()
    search_type = data.get('type', 'song')
    limit = min(int(data.get('limit', 20)), 100)
    platforms = data.get('platforms', [])

    if not query:
        return jsonify({'error': '请输入搜索关键词'}), 400

    if not platforms:
        return jsonify({'error': '请至少选择一个平台'}), 400

    try:
        # 从配置文件获取 API Key
        api_key = current_app.config.get('MUSIC_GATEWAY_API_KEY')
        base_url = current_app.config.get('MUSIC_GATEWAY_BASE_URL')

        # 如果没有 API Key，使用模拟数据
        use_mock = not api_key or api_key == ''

        results = {}

        # QQ 音乐搜索
        if 'qq' in platforms:
            results['qq'] = {
                'platform': 'qq',
                'items': search_qq_music(query, search_type, limit, api_key, base_url, use_mock)
            }

        # 网易云音乐搜索
        if 'netease' in platforms:
            results['netease'] = {
                'platform': 'netease',
                'items': search_netease_music(query, search_type, limit, api_key, base_url, use_mock)
            }

        return jsonify({
            'success': True,
            'results': results,
            'query': query,
            'type': search_type,
            'mock_data': use_mock  # 标识是否使用模拟数据
        })

    except Exception as e:
        import traceback
        print(f"[ERROR] Music search failed: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@api_bp.route('/music/download', methods=['POST'])
@login_required
def download_music():
    """使用网易云音乐 API 下载音乐"""
    data = request.get_json()
    query = data.get('query')          # 歌曲名称（备用）
    provider = data.get('provider')    # 平台: netease, qq
    song_id = data.get('songId')       # 歌曲 ID

    if not song_id and not query:
        return jsonify({'error': '缺少歌曲信息'}), 400

    try:
        from pathlib import Path
        import uuid

        download_dir = Path(__file__).parent.parent / 'music'
        download_dir.mkdir(parents=True, exist_ok=True)

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Referer': 'https://music.163.com'
        }

        # 网易云音乐下载
        if provider == 'netease' and song_id:
            api_url = 'https://music.163.com/api/song/enhance/player/url'
            params = {'id': song_id, 'ids': f'[{song_id}]', 'br': 320000}
            resp = requests.get(api_url, params=params, headers=headers, timeout=10)
            resp.raise_for_status()
            result = resp.json()
            song_url = result.get('data', [{}])[0].get('url')
            if not song_url:
                return jsonify({'success': False, 'error': '无法获取歌曲下载链接'}), 500

        # QQ 音乐下载（暂未实现，保留接口）
        elif provider == 'qq' and song_id:
            return jsonify({'success': False, 'error': 'QQ 音乐下载暂不支持'}), 500

        # 仅有歌名，尝试网易云搜索后下载
        else:
            search_url = 'https://music.163.com/api/search/get'
            search_params = {'s': query, 'type': 1, 'limit': 1}
            resp = requests.get(search_url, params=search_params, headers=headers, timeout=10)
            resp.raise_for_status()
            search_result = resp.json()
            songs = search_result.get('result', {}).get('songs', [])
            if not songs:
                return jsonify({'success': False, 'error': '未找到匹配歌曲'}), 404
            song_id = str(songs[0]['id'])
            api_url = 'https://music.163.com/api/song/enhance/player/url'
            params = {'id': song_id, 'ids': f'[{song_id}]', 'br': 320000}
            resp = requests.get(api_url, params=params, headers=headers, timeout=10)
            resp.raise_for_status()
            result = resp.json()
            song_url = result.get('data', [{}])[0].get('url')
            if not song_url:
                return jsonify({'success': False, 'error': '无法获取歌曲下载链接'}), 500

        # 下载音乐文件
        dl_resp = requests.get(song_url, headers=headers, timeout=60)
        dl_resp.raise_for_status()

        safe_name = f"{query or song_id}_{uuid.uuid4().hex[:8]}.mp3"
        file_path = download_dir / safe_name
        with open(file_path, 'wb') as f:
            f.write(dl_resp.content)

        return jsonify({
            'success': True,
            'message': '下载成功',
            'file': file_path.name,
            'download_url': f'/api/music/file/{file_path.name}'
        })

    except Exception as e:
        import traceback
        print(f"[ERROR] Download failed: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@api_bp.route('/music/lyrics/<song_id>', methods=['GET'])
@login_required
def get_lyrics(song_id):
    """获取歌词"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Referer': 'https://music.163.com'
        }
        resp = requests.get(
            'https://music.163.com/api/song/lyric',
            params={'id': song_id, 'lv': 1},
            headers=headers,
            timeout=10
        )
        resp.raise_for_status()
        data = resp.json()

        lrc = data.get('lrc', {}).get('lyric', '')
        tlyric = data.get('tlyric', {}).get('lyric', '')  # 翻译

        if not lrc:
            return jsonify({'success': False, 'error': '暂无歌词'}), 404

        return jsonify({
            'success': True,
            'lyric': lrc,
            'tlyric': tlyric
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@api_bp.route('/music/file/<filename>', methods=['GET'])
@login_required
def get_music_file(filename):
    """获取下载的音乐文件"""
    try:
        download_dir = Path(__file__).parent.parent / 'music'
        file_path = download_dir / filename

        if not file_path.exists():
            return jsonify({'error': '文件不存在'}), 404

        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def search_qq_music(query, search_type, limit, api_key, base_url, use_mock=False):
    """搜索 QQ 音乐"""
    # 没有 API Key 时，使用网易云音乐 API 提供真实结果
    if use_mock:
        return search_netease_music(query, search_type, limit, api_key, base_url, use_mock=True)

    # 真实 API 调用（需要有效的 API Key）
    try:
        headers = {
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json'
        }

        # QQ 音乐的 provider 名称是 qqmusic
        endpoint = f"{base_url}/v1/qqmusic/search/songs"
        params = {
            'q': query,
            'page': 1,
            'page_size': limit
        }

        response = requests.get(endpoint, headers=headers, params=params, timeout=5)
        response.raise_for_status()

        result = response.json()
        if result.get('code') != 0:
            raise Exception(f"API Error: {result.get('message', 'Unknown error')}")

        songs = result.get('data', {}).get('songs', [])

        # 转换为前端需要的格式
        items = []
        for song in songs:
            items.append({
                'name': song.get('name', '未知歌曲'),
                'artist': ', '.join([a.get('name', '') for a in song.get('singers', [])]) or '未知歌手',
                'album': song.get('album', {}).get('name', ''),
                'duration': song.get('interval', 0),
                'cover': song.get('album', {}).get('pic', ''),
                'url': f"https://y.qq.com/n/ryqq/songDetail/{song.get('mid', '')}" if song.get('mid') else '',
                'id': song.get('mid', '')
            })

        return items

    except Exception as e:
        print(f"[ERROR] QQ Music API failed, using mock data: {str(e)}")
        # API 失败时返回模拟数据
        return search_qq_music(query, search_type, limit, api_key, base_url, use_mock=True)


def search_netease_music(query, search_type, limit, api_key, base_url, use_mock=False):
    """搜索网易云音乐"""
    # 直接使用网易云音乐官方 API（无需 API Key）
    if use_mock:
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Referer': 'https://music.163.com'
            }
            type_map = {'song': 1, 'artist': 100, 'album': 10, 'playlist': 1000}
            search_type_num = type_map.get(search_type, 1)

            resp = requests.get(
                'https://music.163.com/api/search/get',
                params={'s': query, 'type': search_type_num, 'limit': limit},
                headers=headers,
                timeout=10
            )
            resp.raise_for_status()
            result = resp.json()

            if result.get('code') != 200:
                raise Exception(f"API Error: {result.get('message', 'Unknown error')}")

            songs = result.get('result', {}).get('songs', [])

            # 批量获取封面图和歌词可用状态
            items = []
            if songs:
                ids = ','.join([str(s['id']) for s in songs])
                detail_resp = requests.get(
                    'https://music.163.com/api/song/detail',
                    params={'ids': f'[{ids}]'},
                    headers=headers,
                    timeout=10
                )
                detail_resp.raise_for_status()
                detail_songs = detail_resp.json().get('songs', [])
                detail_map = {str(s['id']): s for s in detail_songs}

                for song in songs:
                    sid = str(song.get('id', ''))
                    detail = detail_map.get(sid, {})
                    album = detail.get('album', {})
                    item = {
                        'name': song.get('name', '未知歌曲'),
                        'artist': ', '.join([a.get('name', '') for a in song.get('artists', [])]) or '未知歌手',
                        'album': album.get('name', '') or song.get('album', {}).get('name', ''),
                        'duration': song.get('duration', 0) / 1000,
                        'cover': album.get('picUrl', '') or album.get('blurPicUrl', ''),
                        'url': f"https://music.163.com/#/song?id={sid}" if sid else '',
                        'id': sid
                    }
                    items.append(item)

            return items

        except Exception as e:
            print(f"[ERROR] Netease direct API failed: {str(e)}")
            return [
                {
                    'name': f'{query} - 网易云 {i+1}',
                    'artist': '示例歌手',
                    'album': '示例专辑',
                    'duration': 210 + i * 15,
                    'cover': f'https://via.placeholder.com/150?text=NCM{i+1}',
                    'url': f'https://music.163.com/',
                    'id': f'mock_netease_{i}'
                }
                for i in range(min(5, limit))
            ]

    # 真实网关 API（API Key 有效时使用）
    try:
        headers = {
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json'
        }

        endpoint = f"{base_url}/v1/netease/search/songs"
        params = {
            'q': query,
            'page': 1,
            'page_size': limit
        }

        response = requests.get(endpoint, headers=headers, params=params, timeout=5)
        response.raise_for_status()

        result = response.json()
        if result.get('code') != 0:
            raise Exception(f"API Error: {result.get('message', 'Unknown error')}")

        songs = result.get('data', {}).get('songs', [])

        items = []
        for song in songs:
            items.append({
                'name': song.get('name', '未知歌曲'),
                'artist': ', '.join([a.get('name', '') for a in song.get('artists', [])]) or '未知歌手',
                'album': song.get('album', {}).get('name', ''),
                'duration': song.get('duration', 0) / 1000,
                'cover': song.get('album', ).get('picUrl', ''),
                'url': f"https://music.163.com/#/song?id={song.get('id', '')}" if song.get('id') else '',
                'id': str(song.get('id', ''))
            })

        return items

    except Exception as e:
        print(f"[ERROR] Netease Music API failed, using mock data: {str(e)}")
        return search_netease_music(query, search_type, limit, api_key, base_url, use_mock=True)
